"""
Excel Import/Export Routes
Xuất/nhập dữ liệu Tác giả, Tác phẩm, Nhân vật ra file Excel
"""

import io
import json
from flask import Blueprint, request, jsonify, send_file, session, current_app
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime

excel_bp = Blueprint('excel', __name__, url_prefix='/api/excel')

# ── COLORS ───────────────────────────────────────────────────────────────
BROWN      = "8B4513"
GOLD       = "D4AF37"
DARK       = "3E2723"
CREAM      = "FFF8E1"
LIGHT_GOLD = "FDF6E3"
WHITE      = "FFFFFF"
GREEN      = "E8F5E9"
RED        = "FFEBEE"
BLUE_HDR   = "1F3864"

def neo4j():
    return current_app.config['NEO4J_SERVICE']

def require_admin():
    if not session.get('is_admin'):
        return jsonify({"success": False, "error": "Cần quyền Admin"}), 403
    return None

# ── STYLE HELPERS ─────────────────────────────────────────────────────────

def header_style(cell, bg=BROWN, fg=WHITE, size=12, bold=True):
    cell.font      = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()

def data_style(cell, bg=WHITE, bold=False, center=False):
    cell.font      = Font(name="Arial", size=11, color="2C2C2C", bold=bold)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=True
    )
    cell.border    = thin_border()

def thin_border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def title_row(ws, title, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.font      = Font(bold=True, size=14, color=WHITE, name="Arial")
    c.fill      = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

def info_row(ws, text, ncols, row=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(italic=True, size=10, color="666666", name="Arial")
    c.fill      = PatternFill("solid", fgColor=CREAM)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18

# ═══════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════

@excel_bp.route('/export', methods=['GET'])
def export_excel():
    """Xuất toàn bộ dữ liệu ra 1 file Excel nhiều sheet"""
    err = require_admin()
    if err: return err

    db  = neo4j()
    wb  = Workbook()
    wb.remove(wb.active)  # xóa sheet mặc định

    _build_sheet_tac_gia(wb, db)
    _build_sheet_tac_pham(wb, db)
    _build_sheet_nhan_vat(wb, db)
    _build_sheet_the_loai(wb, db)
    _build_sheet_huong_dan(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"VanHocVN_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname
    )


@excel_bp.route('/export/tac-gia', methods=['GET'])
def export_tac_gia():
    err = require_admin()
    if err: return err
    wb = Workbook(); wb.remove(wb.active)
    _build_sheet_tac_gia(wb, neo4j())
    _build_sheet_huong_dan(wb)
    return _send_wb(wb, "TacGia")


@excel_bp.route('/export/tac-pham', methods=['GET'])
def export_tac_pham():
    err = require_admin()
    if err: return err
    wb = Workbook(); wb.remove(wb.active)
    _build_sheet_tac_pham(wb, neo4j())
    _build_sheet_huong_dan(wb)
    return _send_wb(wb, "TacPham")


def _send_wb(wb, name):
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"VanHocVN_{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname)


# ── SHEET: TÁC GIẢ ─────────────────────────────────────────────────────

def _build_sheet_tac_gia(wb, db):
    ws = wb.create_sheet("Tac Gia")
    ws.freeze_panes = "A4"

    HEADERS = ["Tên tác giả*", "Năm sinh", "Năm mất", "Quê quán",
               "Trường phái", "Bút danh","Giải thưởng", "Tiểu sử",
               "Câu nói nổi tiếng", "Ảnh đại diện (URL)"]
    FIELDS  = ["ten", "nam_sinh", "nam_mat", "que_quan",
               "truong_phai", "but_danh","giai_thuong", "tieu_su",
               "cau_noi_noi_tieng", "anh_dai_dien"]
    WIDTHS  = {chr(65+i): w for i, w in enumerate([28,12,12,20,18,18,22,50,40,40])}

    title_row(ws, "📚 DANH SÁCH TÁC GIẢ - Văn Học Việt Nam", len(HEADERS))
    info_row(ws, f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  (*) bắt buộc", len(HEADERS))

    for ci, h in enumerate(HEADERS, 1):
        header_style(ws.cell(row=3, column=ci, value=h), bg=BROWN)

    data = db.get_all_tac_gia() or []
    for ri, row in enumerate(data, 4):
        bg = LIGHT_GOLD if ri % 2 == 0 else WHITE
        for ci, f in enumerate(FIELDS, 1):
            v = row.get(f, "") or ""
            c = ws.cell(row=ri, column=ci, value=str(v) if v else "")
            data_style(c, bg=bg, center=(ci in [2,3]))

    ws.row_dimensions[3].height = 22
    set_col_widths(ws, WIDTHS)
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"

    # Thêm dòng thống kê cuối
    last = len(data) + 4
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=len(HEADERS))
    c = ws.cell(row=last, column=1, value=f"Tổng cộng: {len(data)} tác giả")
    c.font = Font(bold=True, color=BROWN, italic=True, name="Arial")
    c.alignment = Alignment(horizontal="right")


# ── SHEET: TÁC PHẨM ─────────────────────────────────────────────────────

def _build_sheet_tac_pham(wb, db):
    ws = wb.create_sheet("Tac Pham")
    ws.freeze_panes = "A4"

    HEADERS = ["Tên tác phẩm*", "Tác giả*", "Năm sáng tác", "Năm xuất bản",
               "Thể loại", "Giai đoạn", "Tóm tắt nội dung",
               "Chủ đề chính", "Ý nghĩa", "Giá trị nghệ thuật",
               "Hoàn cảnh sáng tác", "Cấu trúc","Trích đoạn", "Ảnh bìa (URL)"]
    FIELDS  = ["ten", "tac_gia", "nam_sang_tac", "nam_xuat_ban",
               "the_loai", "giai_doan", "noi_dung_tom_tat",
               "chu_de_chinh", "y_nghia", "gia_tri_nghe_thuat",
               "hoan_canh", "cau_truc","trich_doan", "anh_dai_dien"]
    WIDTHS  = {chr(65+i): w for i, w in enumerate([28,22,14,14,16,26,50,40,50,50,50,40,40])}

    title_row(ws, "📖 DANH SÁCH TÁC PHẨM - Văn Học Việt Nam", len(HEADERS))
    info_row(ws, f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  (*) bắt buộc", len(HEADERS))

    for ci, h in enumerate(HEADERS, 1):
        header_style(ws.cell(row=3, column=ci, value=h), bg=BLUE_HDR)

    data = db.get_all_tac_pham(limit=500) or []
    for ri, row in enumerate(data, 4):
        bg = "#EFF6FF" if ri % 2 == 0 else WHITE
        bg = "EFF6FF" if ri % 2 == 0 else WHITE
        for ci, f in enumerate(FIELDS, 1):
            v = row.get(f, "") or ""
            if isinstance(v, list):
                v = "\n\n".join(str(x) for x in v if x)
            c = ws.cell(row=ri, column=ci, value=str(v) if v else "")
            data_style(c, bg=bg, center=(ci in [3,4]))

    ws.row_dimensions[3].height = 22
    set_col_widths(ws, WIDTHS)
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"

    last = len(data) + 4
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=len(HEADERS))
    c = ws.cell(row=last, column=1, value=f"Tổng cộng: {len(data)} tác phẩm")
    c.font = Font(bold=True, color=BLUE_HDR, italic=True, name="Arial")
    c.alignment = Alignment(horizontal="right")


# ── SHEET: NHÂN VẬT ─────────────────────────────────────────────────────

def _build_sheet_nhan_vat(wb, db):
    ws = wb.create_sheet("Nhan Vat")
    ws.freeze_panes = "A4"

    HEADERS = ["Tên nhân vật*", "Tác phẩm*", "Vai trò", "Tính cách",
               "Số phận", "Ý nghĩa điển hình"]
    WIDTHS  = {"A":25, "B":28, "C":18, "D":35, "E":35, "F":40}

    title_row(ws, "👥 DANH SÁCH NHÂN VẬT - Văn Học Việt Nam", len(HEADERS))
    info_row(ws, f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  (*) bắt buộc", len(HEADERS))

    for ci, h in enumerate(HEADERS, 1):
        header_style(ws.cell(row=3, column=ci, value=h), bg="2E7D32")

    query = """
    MATCH (tp:TacPham)-[:CO_NHAN_VAT]->(nv:NhanVat)

    RETURN 
        nv.ten AS ten,
        tp.ten AS tac_pham,

        nv.vai_tro AS vai_tro,
        nv.tinh_cach AS tinh_cach,

        nv.so_phan AS so_phan,
        nv.y_nghia_dien_hinh AS y_nghia_dien_hinh

    ORDER BY tp.ten, nv.ten
    LIMIT 1000
    """
    data = db.execute_query(query) or []

    for ri, row in enumerate(data, 4):
        bg = "F1F8E9" if ri % 2 == 0 else WHITE
        for ci, f in enumerate([
        "ten",
        "tac_pham",
        "vai_tro",
        "tinh_cach",
        "so_phan",
        "y_nghia_dien_hinh"
    ], 1):
            v = row.get(f, "") or ""
            c = ws.cell(row=ri, column=ci, value=str(v) if v else "")
            data_style(c, bg=bg)

    set_col_widths(ws, WIDTHS)
    ws.row_dimensions[3].height = 22
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"

    last = len(data) + 4
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=len(HEADERS))
    c = ws.cell(row=last, column=1, value=f"Tổng cộng: {len(data)} nhân vật")
    c.font = Font(bold=True, color="2E7D32", italic=True, name="Arial")
    c.alignment = Alignment(horizontal="right")


# ── SHEET: THỂ LOẠI ─────────────────────────────────────────────────────

def _build_sheet_the_loai(wb, db):
    ws = wb.create_sheet("The Loai")
    ws.freeze_panes = "A4"

    HEADERS = ["Tên thể loại*", "Mô tả", "Số tác phẩm"]
    WIDTHS  = {"A": 25, "B": 60, "C": 15}

    title_row(ws, "🏷️ DANH SÁCH THỂ LOẠI - Văn Học Việt Nam", len(HEADERS))
    info_row(ws, f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}", len(HEADERS))

    for ci, h in enumerate(HEADERS, 1):
        header_style(ws.cell(row=3, column=ci, value=h), bg="6A1B9A")

    query = """
    MATCH (tl:TheLoai)
    OPTIONAL MATCH (tp:TacPham)-[:THUOC_THE_LOAI]->(tl)
    RETURN tl.ten AS ten, tl.mo_ta AS mo_ta, count(tp) AS so_luong
    ORDER BY so_luong DESC
    """
    data = db.execute_query(query) or []

    for ri, row in enumerate(data, 4):
        bg = "F3E5F5" if ri % 2 == 0 else WHITE
        for ci, (f, ctr) in enumerate([("ten",False),("mo_ta",False),("so_luong",True)], 1):
            v = row.get(f, "") or ""
            c = ws.cell(row=ri, column=ci, value=v)
            data_style(c, bg=bg, center=ctr)

    set_col_widths(ws, WIDTHS)
    ws.row_dimensions[3].height = 22


# ── SHEET: HƯỚNG DẪN IMPORT ─────────────────────────────────────────────

def _build_sheet_huong_dan(wb):
    ws = wb.create_sheet("Huong Dan Import")

    title_row(ws, "📋 HƯỚNG DẪN NHẬP DỮ LIỆU (IMPORT)", 2)

    rows = [
        ("", ""),
        ("🟢 CÁCH NHẬP DỮ LIỆU:", ""),
        ("1.", "Chỉnh sửa trực tiếp trong sheet tương ứng (Tac Gia, Tac Pham...)"),
        ("2.", "Lưu file Excel (.xlsx)"),
        ("3.", "Vào trang Admin → Tab Import/Export → Chọn file → Nhấn 'Nhập Excel'"),
        ("", ""),
        ("📌 LƯU Ý QUAN TRỌNG:", ""),
        ("•", "Cột có dấu (*) là bắt buộc, không được để trống"),
        ("•", "Không xóa hoặc đổi tên các cột tiêu đề (dòng 3)"),
        ("•", "Không xóa dòng tiêu đề và dòng thông tin (dòng 1-3)"),
        ("•", "Tên Tác giả trong sheet Tac Pham phải khớp với sheet Tac Gia"),
        ("•", "Dữ liệu trùng (cùng tên) sẽ được cập nhật, không tạo mới"),
        ("", ""),
        ("🔴 KHÔNG NÊN:", ""),
        ("•", "Thay đổi định dạng cột hoặc thêm cột mới"),
        ("•", "Xóa dòng tiêu đề (3 dòng đầu)"),
        ("•", "Để ký tự đặc biệt trong cột Tên"),
        ("", ""),
        ("📊 ĐỊNH DẠNG DỮ LIỆU:", ""),
        ("Năm:", "Nhập số nguyên, ví dụ: 1820"),
        ("URL ảnh:", "Nhập đường dẫn đầy đủ, ví dụ: https://..."),
        ("Tác phẩm - Thể loại:", "Nhập đúng tên thể loại đã có trong sheet The Loai"),
        ("Tác phẩm - Giai đoạn:", "Ví dụ: Văn học hiện đại (1930-1945)"),
    ]

    for ri, (col_a, col_b) in enumerate(rows, 3):
        ca = ws.cell(row=ri, column=1, value=col_a)
        cb = ws.cell(row=ri, column=2, value=col_b)
        if col_a.startswith(("🟢","📌","🔴","📊")):
            ca.font = Font(bold=True, size=12, color=BROWN, name="Arial")
            ws.row_dimensions[ri].height = 20
        else:
            ca.font = Font(size=11, name="Arial", color="444444")
            cb.font = Font(size=11, name="Arial")
        ca.alignment = Alignment(vertical="center")
        cb.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 70
    ws.sheet_properties.tabColor = "D4AF37"


# ═══════════════════════════════════════════════════════
# IMPORT
# ═══════════════════════════════════════════════════════

@excel_bp.route('/import', methods=['POST'])
def import_excel():
    """Nhập dữ liệu từ file Excel"""
    err = require_admin()
    if err: return err

    if 'file' not in request.files:
        return jsonify({"success": False, "error": "Không có file được gửi lên"}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({"success": False, "error": "Chỉ hỗ trợ file .xlsx"}), 400

    try:
        wb   = load_workbook(file, data_only=True)
        db   = neo4j()
        results = {}

        if "Tac Gia" in wb.sheetnames:
            results["tac_gia"] = _import_sheet_tac_gia(wb["Tac Gia"], db)

        if "Tac Pham" in wb.sheetnames:
            results["tac_pham"] = _import_sheet_tac_pham(wb["Tac Pham"], db)

        if "The Loai" in wb.sheetnames:
            results["the_loai"] = _import_sheet_the_loai(wb["The Loai"], db)

        total_ok  = sum(r.get("success", 0) for r in results.values())
        total_err = sum(r.get("errors",  0) for r in results.values())

        return jsonify({
            "success": True,
            "message": f"Import hoàn tất: {total_ok} bản ghi thành công, {total_err} lỗi",
            "details": results
        })

    except Exception as e:
        return jsonify({"success": False, "error": f"Lỗi đọc file: {str(e)}"}), 500


def _read_sheet_rows(ws, header_row=3):
    """Đọc headers từ dòng header_row, trả về list dict"""
    headers = []
    for cell in ws[header_row]:
        h = str(cell.value or "").strip().rstrip("*")
        headers.append(h)

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        # Bỏ qua dòng tổng kết
        first = str(row[0] or "").strip()
        if first.startswith("Tổng cộng"):
            continue
        d = {}
        for h, v in zip(headers, row):
            d[h] = str(v).strip() if v is not None else ""
        rows.append(d)
    return rows


def _import_sheet_tac_gia(ws, db):
    rows     = _read_sheet_rows(ws)
    ok, errs = 0, 0
    err_list = []

    for r in rows:
        ten = r.get("Tên tác giả", "").strip()
        if not ten:
            errs += 1
            continue
        try:
            # Kiểm tra tồn tại
            existing = db.get_tac_gia_detail(ten)
            data = {
                "ten":              ten,
                "nam_sinh":         _int(r.get("Năm sinh")),
                "nam_mat":          _int(r.get("Năm mất")),
                "que_quan":         r.get("Quê quán", ""),
                "truong_phai":      r.get("Trường phái", ""),
                "but_danh":         r.get("Bút danh", ""),
                "giai_thuong": r.get("Giải thưởng", ""),
                "tieu_su":          r.get("Tiểu sử", ""),
                "cau_noi_noi_tieng":r.get("Câu nói nổi tiếng", ""),
                "anh_dai_dien":     r.get("Ảnh đại diện (URL)", ""),
            }
            if existing:
                query = """
                MATCH (t:TacGia {ten: $ten})
                SET 
                    t.anh_dai_dien = $anh_dai_dien,
                    t.nam_sinh = $nam_sinh,
                    t.but_danh = $but_danh,
                    t.giai_thuong = $giai_thuong,
                    t.cau_noi_noi_tieng = $cau_noi_noi_tieng,
                    t.nam_mat = $nam_mat,
                    t.que_quan = $que_quan,
                    t.truong_phai = $truong_phai,
                    t.tieu_su = $tieu_su
                RETURN t
                """

                db.execute_write(query, data)
            
            else:
                db.create_tac_gia(data)
            ok += 1
        except Exception as e:
            errs += 1
            err_list.append(f"{ten}: {str(e)}")

    return {"success": ok, "errors": errs, "error_list": err_list[:10]}


def _import_sheet_tac_pham(ws, db):
    rows     = _read_sheet_rows(ws)
    ok, errs = 0, 0
    err_list = []

    for r in rows:
        ten = r.get("Tên tác phẩm", "").strip()
        if not ten:
            errs += 1
            continue
        try:
            existing = db.get_tac_pham_detail(ten)
            data = {
                "ten":                ten,
                "tac_gia":            r.get("Tác giả", ""),
                "nam_sang_tac":       _int(r.get("Năm sáng tác")),
                "nam_xuat_ban":       _int(r.get("Năm xuất bản")),
                "the_loai":           r.get("Thể loại", ""),
                "giai_doan":          r.get("Giai đoạn", ""),
                "noi_dung_tom_tat":   r.get("Tóm tắt nội dung", ""),
                "chu_de_chinh":       r.get("Chủ đề chính", ""),
                "y_nghia":            r.get("Ý nghĩa", ""),
                "gia_tri_nghe_thuat": r.get("Giá trị nghệ thuật", ""),
                "hoan_canh":          r.get("Hoàn cảnh sáng tác", ""),
                "cau_truc":           r.get("Cấu trúc", ""),
                "trich_doan": [
                    td.strip()
                    for td in r.get("Trích đoạn", "").split("\n\n")
                    if td.strip()
                ],
                "anh_dai_dien":       r.get("Ảnh bìa (URL)", ""),
                "nhan_vat":           [],
            }
            if existing:
                db.update_tac_pham(ten, data)
            else:
                db.create_tac_pham(data)
            ok += 1
        except Exception as e:
            errs += 1
            err_list.append(f"{ten}: {str(e)}")

    return {"success": ok, "errors": errs, "error_list": err_list[:10]}


def _import_sheet_the_loai(ws, db):
    rows     = _read_sheet_rows(ws)
    ok, errs = 0, 0

    for r in rows:
        ten = r.get("Tên thể loại", "").strip()
        if not ten:
            errs += 1
            continue
        try:
            # Upsert TheLoai
            db.execute_query(
                "MERGE (tl:TheLoai {ten: $ten}) SET tl.mo_ta = $mo_ta",
                {"ten": ten, "mo_ta": r.get("Mô tả", "")}
            )
            ok += 1
        except Exception:
            errs += 1

    return {"success": ok, "errors": errs}


def _int(val):
    try:
        return int(float(str(val))) if val and str(val).strip() not in ("", "None") else None
    except:
        return None


# ── DOWNLOAD TEMPLATE ────────────────────────────────────────────────────

@excel_bp.route('/template', methods=['GET'])
def download_template():
    """Tải file template Excel rỗng để nhập dữ liệu mới"""
    err = require_admin()
    if err: return err

    wb = Workbook(); wb.remove(wb.active)

    # Tạo sheet template rỗng với header
    def make_template_sheet(wb, sheet_name, headers, title, bg):
        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = "A4"
        title_row(ws, title, len(headers))
        info_row(ws, "Nhập dữ liệu từ dòng 4 trở xuống. Cột có (*) là bắt buộc.", len(headers))
        for ci, h in enumerate(headers, 1):
            header_style(ws.cell(row=3, column=ci, value=h), bg=bg)
        # 5 dòng mẫu trống
        for ri in range(4, 9):
            bg_r = LIGHT_GOLD if ri % 2 == 0 else WHITE
            for ci in range(1, len(headers)+1):
                data_style(ws.cell(row=ri, column=ci, value=""), bg=bg_r)
        ws.row_dimensions[3].height = 22
        return ws

    make_template_sheet(wb, "Tac Gia",
        ["Tên tác giả*","Năm sinh","Năm mất","Quê quán","Trường phái",
         "Bút danh","Tiểu sử","Câu nói nổi tiếng","Ảnh đại diện (URL)"],
        "📚 NHẬP TÁC GIẢ MỚI", BROWN)

    make_template_sheet(wb, "Tac Pham",
        ["Tên tác phẩm*","Tác giả*","Năm sáng tác","Năm xuất bản",
         "Thể loại","Giai đoạn","Tóm tắt nội dung","Chủ đề chính",
         "Ý nghĩa","Giá trị nghệ thuật","Hoàn cảnh sáng tác","Cấu trúc","Ảnh bìa (URL)"],
        "📖 NHẬP TÁC PHẨM MỚI", BLUE_HDR)

    make_template_sheet(wb, "The Loai",
        ["Tên thể loại*","Mô tả"],
        "🏷️ NHẬP THỂ LOẠI MỚI", "6A1B9A")

    _build_sheet_huong_dan(wb)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="VanHocVN_Template_Import.xlsx")